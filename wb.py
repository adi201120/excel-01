import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "adityaA": {
        "math": 65,
        "science": 78,
        "english": 98,
        "python": 89
    },
    "adityaU": {
        "math": 55,
        "science": 72,
        "english": 87,
        "python": 95
    },
    "prathamesh": {
        "math": 100,
        "science": 45,
        "english": 75,
        "python": 92
    },
    "bunny": {
        "math": 30,
        "science": 25,
        "english": 45,
        "python": 100
    },
    "prahlad": {
        "math": 100,
        "science": 100,
        "english": 100,
        "python": 60
    }
}

wb = Workbook()
ws = wb.active
ws.title = "test01"

headings = ['Name'] + list(data['adityaA'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

for col in range(2, len(data['adityaA']) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("Newfile.xlsx")

